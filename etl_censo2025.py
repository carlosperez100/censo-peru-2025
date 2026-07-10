# -*- coding: utf-8 -*-
"""
ETL Censo 2025 (INEI) - Tabulados de Poblacion  ->  base relacional unica (SQLite) + grafo
Autor: GEMSES / Carlos Perez.  Metodo: ingesta universal sin perdida (celdas combinadas
desanidadas a formato largo) + dimension geografica que enlaza por nombre normalizado.

Entrada : ./raw/*.xlsx  (8 temas descargados del portal INEI)
Salida  : ./censo2025.db (SQLite)  +  ./grafo/ubigeo.json  +  ./DICCIONARIO.md
"""
import openpyxl, sqlite3, glob, os, re, json, csv, unicodedata, difflib

RAW = os.path.join(os.path.dirname(__file__), "raw")
DB  = os.path.join(os.path.dirname(__file__), "censo2025.db")
GRAFO = os.path.join(os.path.dirname(__file__), "grafo")
UBIGEO_CSV = os.path.join(os.path.dirname(__file__), "ref", "ubigeo_inei.csv")
os.makedirs(GRAFO, exist_ok=True)

TEMAS = {  # archivo -> (codigo_tema, nombre_tema)
 "01_indicadores_demograficos.xlsx": ("DEM", "Indicadores demograficos"),
 "02_fecundidad.xlsx":               ("FEC", "Fecundidad"),
 "03_migracion.xlsx":                ("MIG", "Migracion"),
 "04_estado_civil_identidad_seguro.xlsx": ("ECS", "Estado civil, identidad y seguro de salud"),
 "05_educacion.xlsx":                ("EDU", "Educacion"),
 "06_discapacidad.xlsx":             ("DIS", "Discapacidad"),
 "07_etnicidad.xlsx":                ("ETN", "Etnicidad"),
 "08_poblacion_edad_trabajar.xlsx":  ("PET", "Poblacion en edad de trabajar"),
 "09_vivienda_caracteristicas.xlsx": ("VIV", "Vivienda - caracteristicas"),
 "10_vivienda_servicios.xlsx":       ("VSB", "Vivienda - servicios basicos"),
}

def norm(s):
    """normaliza texto: minusculas, sin tildes, sin espacios extra, sin marca de nota (2/)."""
    if s is None: return ""
    s = str(s).replace("\n", " ").strip()
    s = re.sub(r"\s*\d+/\s*$", "", s)          # quita marcador de nota al pie: "Lima 2/" -> "Lima"
    s = " ".join(s.split())
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s.lower()

def clean(s):
    if s is None: return ""
    return " ".join(str(s).replace("\n", " ").split()).strip()

def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)

def build_grid(ws):
    """Devuelve grid (lista de listas 1-indexed) con celdas combinadas rellenadas."""
    maxr, maxc = ws.max_row, ws.max_column
    grid = [[None]*(maxc+1) for _ in range(maxr+1)]
    for row in ws.iter_rows():
        for c in row:
            grid[c.row][c.column] = c.value
    for mr in ws.merged_cells.ranges:
        v = grid[mr.min_row][mr.min_col]
        for r in range(mr.min_row, mr.max_row+1):
            for c in range(mr.min_col, mr.max_col+1):
                if grid[r][c] is None:
                    grid[r][c] = v
    return grid, maxr, maxc

def parse_sheet(grid, maxr, maxc):
    """Extrae titulo, etiqueta de la columna-fila, rutas de columnas y filas de datos."""
    titulo = ""
    for c in range(1, maxc+1):
        if grid[1][c]:
            titulo = clean(grid[1][c]); break
    # primera fila de datos = primera fila (>=2) con algun numero en col>=2
    data_start = None
    for r in range(2, maxr+1):
        if any(is_num(grid[r][c]) for c in range(2, maxc+1)):
            data_start = r; break
    if data_start is None:
        return titulo, "", {}, []
    header_rows = [r for r in range(2, data_start) if any(clean(grid[r][c]) for c in range(1, maxc+1))]
    dim_fila = " / ".join(dict.fromkeys(clean(grid[r][1]) for r in header_rows if clean(grid[r][1])))
    # ruta de cada columna de datos (concatena tokens de cabecera de arriba a abajo)
    col_path = {}
    for c in range(2, maxc+1):
        toks = []
        for r in header_rows:
            t = clean(grid[r][c])
            if t and (not toks or toks[-1] != t):
                toks.append(t)
        col_path[c] = " | ".join(toks)
    # unicidad posicional: si dos columnas comparten cabecera dentro del MISMO cuadro
    # (p.ej. bandas Total/Urbana/Rural que el Excel rotula igual), se sufija #2, #3...
    vistos = {}
    for c in sorted(col_path):
        p = col_path[c] or "(sin cabecera)"
        vistos[p] = vistos.get(p, 0) + 1
        col_path[c] = p if vistos[p] == 1 else f"{p} #{vistos[p]}"
    # filas de datos
    filas = []
    for r in range(data_start, maxr+1):
        etq = clean(grid[r][1])
        vals = [(c, grid[r][c]) for c in range(2, maxc+1) if is_num(grid[r][c])]
        if not etq and not vals:
            continue
        filas.append((r, etq, vals))
    return titulo, dim_fila, col_path, filas

# ---- deteccion de unidades geograficas (jerarquia INEI: PERU > DEPARTAMENTO > PROVINCIA > DISTRITO)
DEPARTAMENTOS = {  # nombre_norm -> codigo ubigeo depto (2 digitos, estandar INEI)
 "amazonas":"01","ancash":"02","apurimac":"03","arequipa":"04","ayacucho":"05","cajamarca":"06",
 "callao":"07","cusco":"08","huancavelica":"09","huanuco":"10","ica":"11","junin":"12",
 "la libertad":"13","lambayeque":"14","lima":"15","loreto":"16","madre de dios":"17","moquegua":"18",
 "pasco":"19","piura":"20","puno":"21","san martin":"22","tacna":"23","tumbes":"24","ucayali":"25",
}   # Callao / Lima Metropolitana / Region Lima se resuelven en cod_top() con codigos propios

def cod_top(nombre_norm, orden):
    """Codigo de unidad territorial de primer nivel (departamento/Callao/Lima)."""
    if nombre_norm in DEPARTAMENTOS:
        return DEPARTAMENTOS[nombre_norm] + "00"
    if "callao" in nombre_norm:               return "0700"
    if "lima metropolitana" in nombre_norm:   return "1500"
    if "region lima" in nombre_norm:          return "15R0"
    return f"{orden:02d}00"                    # respaldo

def build_ubigeo(archivo_flagship):
    """Construye dim_ubigeo desde el cuadro insignia usando los ENCABEZADOS DE BLOQUE
    (la etiqueta que precede a un bloque de edades ES una unidad geografica)."""
    wb = openpyxl.load_workbook(os.path.join(RAW, archivo_flagship), data_only=True)
    ws = wb["INDDEM01"]
    grid, maxr, maxc = build_grid(ws)
    _, _, _, filas = parse_sheet(grid, maxr, maxc)
    # grupos de edad: "menores de..." o contienen "año/años" como palabra.
    # OJO: usar limite de palabra (\bano\b), si no "Marañon"->"maranon" contiene "ano".
    def es_edad(n):
        return n.startswith("menores") or bool(re.search(r"\banos?\b", n))
    # geografia = etiquetas que NO son grupos de edad ni notas/fuentes al pie
    def es_ruido(n):
        return n.startswith("nota") or n.startswith("fuente") or n.startswith("1/") or len(n) > 55
    heads = [clean(etq) for _, etq, vals in filas
             if clean(etq) and not es_edad(norm(etq)) and not es_ruido(norm(etq))]
    peru_id = "0000"
    ubigeos = [(peru_id, "Peru", "peru", "nacional", None, None)]
    seen = {peru_id}
    cur_top = peru_id; cur_prov = None
    prov_ctr = {}; dist_ctr = {}
    for h in heads:
        n = norm(h)
        if n == "peru":
            cur_top = peru_id; cur_prov = None
        elif n.startswith("provincia "):                     # --- PROVINCIA ---
            nombre = clean(h)[len("PROVINCIA "):].strip()
            cod = cur_top[:2]
            prov_ctr[cod] = prov_ctr.get(cod, 0) + 1
            uid = cod + f"{prov_ctr[cod]:02d}"
            while uid in seen: uid += "x"
            ubigeos.append((uid, nombre.title(), norm(nombre), "provincia", cur_top, cod))
            seen.add(uid); cur_prov = uid; dist_ctr[uid] = 0
        elif n.startswith("distrito "):                      # --- DISTRITO ---
            nombre = clean(h)[len("DISTRITO "):].strip()
            padre = cur_prov or cur_top
            dist_ctr[padre] = dist_ctr.get(padre, 0) + 1
            uid = padre + f"d{dist_ctr[padre]:02d}"
            while uid in seen: uid += "x"
            ubigeos.append((uid, nombre.title(), norm(nombre), "distrito", padre, cur_top[:2]))
            seen.add(uid)
        else:                                                # --- DEPARTAMENTO / Callao / Lima ---
            uid = cod_top(n, len([u for u in ubigeos if u[3]=="departamento"])+1)
            nombre = "Prov. Const. del Callao" if "callao" in n else clean(h).title()
            if uid not in seen:
                ubigeos.append((uid, nombre, n, "departamento", peru_id, uid[:2]))
                seen.add(uid); prov_ctr[uid[:2]] = 0
            cur_top = uid; cur_prov = None
    wb.close()
    return ubigeos

# ---------------------------------------------------------------- indexado UBIGEO oficial
ALIAS_GEO = {"nasca": "nazca", "kimbiri": "quimbiri", "raimondi": "raymondi"}  # censo 2025 -> catalogo

def g(n):
    """Normalizacion geografica extra: quita parentesis y aplica alias de grafia."""
    n = re.sub(r"\(.*?\)", "", n)
    n = " ".join(n.replace("-", " ").split())
    return ALIAS_GEO.get(n, n)

def cargar_catalogo():
    """Lee ref/ubigeo_inei.csv -> lookups por nombre normalizado (tripleta jerarquica)."""
    dist_lk, prov_lk, dep_lk = {}, {}, {}
    prov_dist = {}   # DDPP -> [(dist_norm, inei)]  (para fuzzy dentro de la provincia)
    with open(UBIGEO_CSV, encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            u = row["inei"].strip()
            de, pr, di = g(norm(row["departamento"])), g(norm(row["provincia"])), g(norm(row["distrito"]))
            dist_lk[(de, pr, di)] = u
            prov_lk.setdefault((de, pr), u[:4] + "00")
            dep_lk.setdefault(de, u[:2] + "0000")
            prov_dist.setdefault(u[:4], []).append((di, u))
    # overrides manuales para distritos nuevos/renombrados (ref/ubigeo_overrides.csv)
    ov = os.path.join(os.path.dirname(UBIGEO_CSV), "ubigeo_overrides.csv")
    if os.path.exists(ov):
        with open(ov, encoding="utf-8") as fh:
            for row in csv.DictReader(fh):
                if not row.get("inei", "").strip(): continue
                dist_lk[(g(norm(row["departamento"])), g(norm(row["provincia"])), g(norm(row["distrito"])))] = row["inei"].strip()
    return dist_lk, prov_lk, dep_lk, prov_dist

def _dep_cat(n):
    """Nombre de departamento del nodo -> departamento del catalogo oficial."""
    if "callao" in n: return "callao"
    if "lima metropolitana" in n or "region lima" in n: return "lima"
    return n

def resolver_ubigeo(ubigeos):
    """Devuelve {ubigeo_id -> codigo INEI oficial (6 digitos)} emparejando por jerarquia.
    Tiers: exacto -> alias -> override -> fuzzy dentro de la MISMA provincia."""
    dist_lk, prov_lk, dep_lk, prov_dist = cargar_catalogo()
    byid = {u[0]: u for u in ubigeos}            # (uid, nombre, norm, nivel, padre, cod_dep)
    out = {}
    # ---- Pase 1: exacto / alias / override / departamento / provincia (sin fuzzy)
    def contexto_dist(padre):
        pa = byid.get(padre)
        if pa and pa[3] == "provincia":
            gp = byid.get(pa[4]); return (_dep_cat(gp[2]) if gp else ""), pa[2]
        if pa and "callao" in pa[2]:             return "callao", "callao"
        if pa and "lima metropolitana" in pa[2]: return "lima", "lima"
        return (_dep_cat(pa[2]), pa[2]) if pa else ("", "")
    for uid, nom, n, nivel, padre, cod in ubigeos:
        code = None
        if nivel == "nacional":
            code = "000000"
        elif nivel == "departamento":
            if "callao" in n:              code = "070000"
            elif "lima metropolitana" in n: code = "150100"
            elif "region lima" in n:        code = "150000"
            else:                           code = dep_lk.get(n)
        elif nivel == "provincia":
            pa = byid.get(padre); de = _dep_cat(pa[2]) if pa else ""
            code = prov_lk.get((g(de), g(n)))
        elif nivel == "distrito":
            de, pr = contexto_dist(padre)
            code = dist_lk.get((g(de), g(pr), g(n)))
        out[uid] = code
    usados = {c for c in out.values() if c}
    # ---- Pase 2: fuzzy dentro de la MISMA provincia, umbral alto y SIN duplicar codigo
    fuzzy_log = []
    for uid, nom, n, nivel, padre, cod in ubigeos:
        if out[uid] or nivel != "distrito":
            continue
        de, pr = contexto_dist(padre)
        ddpp = (prov_lk.get((g(de), g(pr))) or "")[:4]
        best, score = None, 0.0
        for dn, du in prov_dist.get(ddpp, []):
            if du in usados:            # no reasignar un codigo ya tomado
                continue
            r = difflib.SequenceMatcher(None, g(n), dn).ratio()
            if r > score: best, score = du, r
        if best and score >= 0.85:
            out[uid] = best; usados.add(best); fuzzy_log.append((nom, best, round(score, 2)))
    if fuzzy_log:
        print("  fuzzy aceptado (>=0.85, provincia, sin dup): " +
              "; ".join(f"{a}->{b}({s})" for a, b, s in fuzzy_log))
    return out

# ---------------------------------------------------------------- carga a SQLite
def main():
    if os.path.exists(DB): os.remove(DB)
    con = sqlite3.connect(DB); cx = con.cursor()
    cx.executescript("""
    CREATE TABLE tema(cod_tema TEXT PRIMARY KEY, nombre TEXT, archivo TEXT, peso_kb INTEGER);
    CREATE TABLE cuadro(cuadro_id INTEGER PRIMARY KEY, cod_tema TEXT, hoja TEXT, titulo TEXT,
                        dimension_fila TEXT, n_datos INTEGER,
                        FOREIGN KEY(cod_tema) REFERENCES tema(cod_tema));
    CREATE TABLE nota(cod_tema TEXT, texto TEXT);
    CREATE TABLE dim_ubigeo(ubigeo_id TEXT PRIMARY KEY, nombre TEXT, nombre_norm TEXT,
                            nivel TEXT, padre_id TEXT, cod_departamento TEXT,
                            ubigeo_inei TEXT);
    -- dimensiones factorizadas (star schema) -> base ligera
    CREATE TABLE dim_fila(fila_id INTEGER PRIMARY KEY, etiqueta TEXT, etiqueta_norm TEXT);
    CREATE TABLE dim_columna(col_id INTEGER PRIMARY KEY, columna TEXT);
    CREATE TABLE dato(
        dato_id INTEGER PRIMARY KEY, cuadro_id INTEGER, fila_orden INTEGER,
        fila_id INTEGER, col_id INTEGER, valor REAL,
        FOREIGN KEY(cuadro_id) REFERENCES cuadro(cuadro_id),
        FOREIGN KEY(fila_id)   REFERENCES dim_fila(fila_id),
        FOREIGN KEY(col_id)    REFERENCES dim_columna(col_id));
    """)

    # dim_ubigeo + indexado UBIGEO oficial (INEI)
    ubis = build_ubigeo("01_indicadores_demograficos.xlsx")
    inei = resolver_ubigeo(ubis)
    ubis_i = [u + (inei.get(u[0]),) for u in ubis]
    cx.executemany("INSERT INTO dim_ubigeo VALUES(?,?,?,?,?,?,?)", ubis_i)
    ubi_norms = {u[2] for u in ubis}
    n_ok = sum(1 for u in ubis if inei.get(u[0]))
    print(f"  UBIGEO INEI resuelto: {n_ok}/{len(ubis)} nodos")
    # exportar pendientes (nodos sin codigo) con su contexto -> plantilla para overrides
    byid = {u[0]: u for u in ubis}
    def contexto(u):
        dep = prov = ""
        if u[3] == "provincia":
            pa = byid.get(u[4]); dep = pa[1] if pa else ""; prov = u[1]
        elif u[3] == "distrito":
            pa = byid.get(u[4])
            if pa and pa[3] == "provincia":
                gp = byid.get(pa[4]); dep = gp[1] if gp else ""; prov = pa[1]
            elif pa: dep = pa[1]; prov = pa[1]
        return dep, prov
    pend = [u for u in ubis if not inei.get(u[0]) and u[3] != "nacional"]
    with open(os.path.join(os.path.dirname(UBIGEO_CSV), "ubigeo_pendientes.csv"), "w",
              encoding="utf-8", newline="") as fh:
        w = csv.writer(fh); w.writerow(["ubigeo_id", "nivel", "departamento", "provincia", "distrito", "inei"])
        for u in pend:
            dep, prov = contexto(u)
            w.writerow([u[0], u[3], dep, prov, u[1] if u[3] == "distrito" else "", ""])

    fila_cache = {}; col_cache = {}   # texto -> id  (deduplicacion)
    def fid(etq, en):
        k = etq
        if k not in fila_cache:
            i = len(fila_cache) + 1
            fila_cache[k] = i
            cx.execute("INSERT INTO dim_fila VALUES(?,?,?)", (i, etq, en))
        return fila_cache[k]
    def cid(col):
        if col not in col_cache:
            i = len(col_cache) + 1
            col_cache[col] = i
            cx.execute("INSERT INTO dim_columna VALUES(?,?)", (i, col))
        return col_cache[col]

    cuadro_id = 0; dato_id = 0
    for archivo, (cod_tema, nombre_tema) in TEMAS.items():
        path = os.path.join(RAW, archivo)
        if not os.path.exists(path):
            print("FALTA:", archivo); continue
        peso = os.path.getsize(path)//1024
        cx.execute("INSERT INTO tema VALUES(?,?,?,?)", (cod_tema, nombre_tema, archivo, peso))
        wb = openpyxl.load_workbook(path, data_only=True)
        for hoja in wb.sheetnames:
            ws = wb[hoja]
            if norm(hoja).startswith("nota de presentacion"):
                txt = " ".join(clean(c.value) for row in ws.iter_rows() for c in row if c.value)
                cx.execute("INSERT INTO nota VALUES(?,?)", (cod_tema, txt[:8000]))
                continue
            grid, maxr, maxc = build_grid(ws)
            titulo, dim_fila, col_path, filas = parse_sheet(grid, maxr, maxc)
            cuadro_id += 1
            n_datos = 0
            batch = []
            for r, etq, vals in filas:
                f_id = fid(etq, norm(etq))
                for c, v in vals:
                    dato_id += 1; n_datos += 1
                    batch.append((dato_id, cuadro_id, r, f_id, cid(col_path.get(c, "")), float(v)))
            cx.executemany("INSERT INTO dato VALUES(?,?,?,?,?,?)", batch)
            cx.execute("INSERT INTO cuadro VALUES(?,?,?,?,?,?)",
                       (cuadro_id, cod_tema, hoja, titulo, dim_fila, n_datos))
        wb.close()
        print(f"OK tema {cod_tema}: {archivo}")

    # vistas utiles
    cx.executescript("""
    CREATE VIEW v_dato AS
      SELECT d.dato_id, t.cod_tema, t.nombre AS tema, c.hoja, c.titulo,
             f.etiqueta AS etiqueta_fila, f.etiqueta_norm, k.columna, d.valor
      FROM dato d JOIN cuadro c   ON c.cuadro_id=d.cuadro_id
                  JOIN tema t     ON t.cod_tema=c.cod_tema
                  JOIN dim_fila f ON f.fila_id=d.fila_id
                  JOIN dim_columna k ON k.col_id=d.col_id;
    -- datos con geografia resuelta: JOIN por nombre normalizado (sin tildes) Y por nivel
    -- inferido del prefijo de la etiqueta. Nivel nacional/departamento/provincia (los
    -- nombres son unicos); a nivel distrito hay homonimos -> requiere codigo ubigeo oficial.
    CREATE VIEW v_dato_geo AS
      SELECT d.*, u.ubigeo_id, u.ubigeo_inei, u.nombre AS ubigeo_nombre, u.nivel AS ubigeo_nivel,
             u.cod_departamento
      FROM v_dato d JOIN dim_ubigeo u
        ON u.nombre_norm = TRIM(CASE
             WHEN d.etiqueta_norm LIKE 'provincia %' THEN SUBSTR(d.etiqueta_norm,11)
             ELSE d.etiqueta_norm END)
       AND u.nivel = CASE
             WHEN d.etiqueta_norm LIKE 'provincia %' THEN 'provincia'
             WHEN d.etiqueta_norm = 'peru'           THEN 'nacional'
             ELSE 'departamento' END;
    -- AFILIACION A SEGURO DE SALUD por ubigeo (SALUD06) -> \"donde estan los afiliados\"
    CREATE VIEW v_seguro AS
      SELECT ubigeo_id, ubigeo_inei, ubigeo_nombre, ubigeo_nivel, cod_departamento,
        SUM(CASE WHEN columna='Total'                     THEN valor END) AS poblacion,
        SUM(CASE WHEN columna='Ninguno'                   THEN valor END) AS sin_seguro,
        SUM(CASE WHEN columna LIKE '%Seguro Integral%'    THEN valor END) AS sis,
        SUM(CASE WHEN columna LIKE '%EsSalud%'            THEN valor END) AS essalud,
        SUM(CASE WHEN columna LIKE '%fuerzas armadas%'    THEN valor END) AS ffaa_pnp,
        SUM(CASE WHEN columna LIKE '%privado%'            THEN valor END) AS privado,
        SUM(CASE WHEN columna LIKE '%Otro seguro%'        THEN valor END) AS otro
      FROM v_dato_geo WHERE hoja='SALUD06'
      GROUP BY ubigeo_id, ubigeo_inei, ubigeo_nombre, ubigeo_nivel, cod_departamento;
    """)
    # indices
    cx.executescript("""
    CREATE INDEX ix_dato_cuadro ON dato(cuadro_id);
    CREATE INDEX ix_dato_fila   ON dato(fila_id);
    CREATE INDEX ix_fila_norm   ON dim_fila(etiqueta_norm);
    CREATE INDEX ix_ubi_norm    ON dim_ubigeo(nombre_norm);
    CREATE INDEX ix_ubi_inei    ON dim_ubigeo(ubigeo_inei);
    """)
    con.commit()
    cx.execute("VACUUM"); con.commit()

    # ---- exportar grafo (jerarquia ubigeo) para graphify / gemses-grafos
    nodos = [{"id":u[0],"label":u[3],"nombre":u[1],"nivel":u[3],"cod_dep":u[5]} for u in ubis]
    aristas = [{"origen":u[4],"destino":u[0],"tipo":"CONTIENE"} for u in ubis if u[4]]
    json.dump({"nodos":nodos,"aristas":aristas}, open(os.path.join(GRAFO,"ubigeo.json"),"w",encoding="utf8"),
              ensure_ascii=False, indent=2)

    # resumen
    for q,lbl in [("SELECT COUNT(*) FROM tema","temas"),("SELECT COUNT(*) FROM cuadro","cuadros"),
                  ("SELECT COUNT(*) FROM dato","datos"),("SELECT COUNT(*) FROM dim_ubigeo","ubigeos"),
                  ("SELECT COUNT(*) FROM v_dato_geo","datos_geo")]:
        print(f"  {lbl:>10}: {cx.execute(q).fetchone()[0]:,}")
    con.close()

    # ---- integrar OFERTA DE SALUD de SUSALUD (IPRESS + RRHH + densidad por 10k) ----
    if os.path.exists(os.path.join(RAW, "susalud", "RENIPRESS.csv")):
        try:
            import integrar_salud
            integrar_salud.main(DB)
            print("  SUSALUD integrado (ipress + salud_rrhh + densidad_salud)")
        except Exception as e:
            print("  (SUSALUD no integrado:", e, ")")

    print("Peso censo2025.db:", os.path.getsize(DB)//1024, "KB")

if __name__ == "__main__":
    main()
